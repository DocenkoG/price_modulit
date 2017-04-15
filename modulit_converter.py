# -*- coding: UTF-8 -*-
import os
import os.path
import logging
import logging.config
import io
import sys
import configparser
import time
import openpyxl                        # Для .xlsx
import xlrd                            # для .xls


def make_loger():
    global log
    logging.config.fileConfig('logging.cfg')
    log = logging.getLogger('logFile')



def convert2csv( myname ):
    global log
    global SheetName
    global FilenameIn
    global FilenameOut
    global out_columns_names
    global out_columns_j
    global in_columns_j
    global colGrp
    global colSGrp
    global GrpFonti
    global BrandFonti
    global SubGrpFonti
    global HeaderFontSize
    global RegularFontSize
    global SubGrpBackgroundColor
    global GrpBackgroundColor
    global strHeader
    global SubGrpFontSize
    global GrpFontSize
    make_loger()
    log.debug('Begin ' + __name__ + 'convert2csv')

    # Прочитать конфигурацию из файла
    ff = config_read( myname )
    log.debug('Открываю файл '+ FilenameIn)
    book = openpyxl.load_workbook(filename = FilenameIn, read_only=False, keep_vba=False, data_only=False)
#    book = xlrd.open_workbook( FilenameIn.encode('cp1251'), formatting_info=False)
#    book = xlrd.open_workbook( os.path.join( mydir, FilenameIn.encode('cp1251')), formatting_info=True)
    
    log.debug('Устанавливаю страницу ' + SheetName )
    sh = book[SheetName]

    ssss = []
                                                        # цикл по строкам файла
    print( 'На странице %d строк' % book[SheetName].max_row)
    for i in range(book[SheetName].min_row, book[SheetName].max_row+1) :
        i_last = i
        try:
            ccc = sh.cell(row=i, column=colGrp)
            if ccc.value == None :
                print (i, colGrp, 'Пусто!!!')
                continue
            '''                                        # Атрибуты шрифта для настройки конфига
            print( 'Строка', i, ccc.value,)
            print( 'font=',   ccc.font.name,)
            print( 'bold=',   ccc.font.bold,)
            print( 'italic=', ccc.font.italic,)
            print( 'size=',   ccc.font.size)
            #print( 'colour=', ccc.font.color.rgb)
            print( 'background=',ccc.fill.fill_type)
            print( 'backgroundColor1=', ccc.fill.start_color)
            print( 'backgroundColor2=', ccc.fill.end_color)
            print( 'Строка', i, 'столбец', colGrp, 'значение', ccc.value)
            continue
            '''

            if  GrpFontSize <= ccc.font.size :                      # Группа
                grpName = quoted(sh.cell(row=i, column=colGrp).value)
                subGrpName = ''
                print('группа', grpName)
    
#            elif SubGrpFontSize == ccc.font.size :                  # Подгруппа
#                subGrpName = quoted(sh.cell(row=i,column=colSGrp).value)
    
            elif HeaderFontSize == ccc.font.size :                            # Заголовок таблицы
                pass
    
            elif (None == sh.cell(row=i, column=in_columns_j['закупка']).value) :    # Пустая строка
                pass
                print( 'Пустая строка:', sh.cell(row=i, column=in_columns_j['закупка']).value )
    
            elif RegularFontSize == ccc.font.size :                 # Информационная строка
                ccc = sh.cell(row=i, column=out_columns_j['код'])
                code = ccc.value
                sss = []                                    # формируемая строка для вывода в файл
                for strname in out_columns_names :
                    if strname in out_columns_j :
                        # берем значение из соответствующей ячейки файла
                        j = out_columns_j[strname] 
                        ccc = sh.cell(row=i, column=j)
                        cellType  = ccc.data_type
                        cellValue = ccc.value
#                       print (cellType, cellValue)
                        if cellValue == None : 
                            ss = ''
                        elif cellType in ('n') :                            # numeric
                            if int(cellValue) == cellValue:
                                ss = str(int(cellValue))
                            else :
                                ss = str(cellValue)
                        elif strname in ('закупка','продажа','цена1', 'цена2') :
                            ss = '0' 
                        elif cellType == 's' :
                            ss = quoted(cellValue ) 
                        else:
                            ss = ''
                    else : 
                        # вычисляемое поле
                        s1 = sh.cell(row=i, column=in_columns_j['бренд']).value
                        s2 = sh.cell(row=i, column=in_columns_j['код']).value
                        s3 = sh.cell(row=i, column=in_columns_j['примечание']).value
                        ss = quoted(s1 + ' ' + s2 + ' ' + s3)
                        pass
                    sss.append(ss)
    
                sss.append(grpName)
                sss.append(subGrpName)
                ssss.append(','.join(sss))
            else :
                loger.debug('Нераспознана строка: <' + sh.cell(row=i, column=out_columns_j['код']).value + '>' )
        except Exception as e:
            loger.debug('Exception: <' + str(e) + '> при обработке строки ' + str(i) +'<' + '>' )
            raise e

    
    f2 = open( FilenameOut, 'w', encoding='cp1251')
    f2.write(strHeader  + ',\n')
    data = ',\n'.join(ssss) +','
    dddd = data.encode(encoding='cp1251', errors='replace')
    data = dddd.decode(encoding='cp1251')
    f2.write(data)
    f2.close()

'''
            else :
                loger.debug('Нераспознана строка: <' + sh.cell(row=i, column=out_columns_j['код']).value + '>' )
        except Exception as e:
            loger.debug('Exception: <' + str(e) + '> при обработке строки ' + str(i) +'<' + '>' )
            raise e
'''    



def config_read( myname ):
    global log
    global SheetName
    global FilenameIn
    global FilenameOut
    global out_columns_names
    global out_columns_j
    global in_columns_j
    global colGrp
    global colSGrp
    global GrpFonti
    global SubGrpFonti
    global BrandFonti
    global HeaderFontSize
    global RegularFontSize
    global SubGrpBackgroundColor
    global GrpBackgroundColor
    global strHeader
    global SubGrpFontSize
    global GrpFontSize

    cfgFName = myname + '.cfg'
    log.debug('Begin config_read ' + cfgFName )
    
    config = configparser.ConfigParser()
    if os.path.exists(cfgFName):     config.read( cfgFName)
    else : log.debug('Не найден файл конфигурации.')

    # в разделе [cols_in] находится список интересующих нас колонок и номера столбцов исходного файла
    in_columns_names = config.options('cols_in')
    in_columns_j = {}
    for vName in in_columns_names :
        if ('' != config.get('cols_in', vName)) :
            in_columns_j[vName] = config.getint('cols_in', vName) 
    
    # По разделу [cols_out] формируем перечень выводимых колонок и строку заголовка результирующего CSV файла
    temp_list = config.options('cols_out')
    temp_list.sort()

    out_columns_names = []
    for vName in temp_list :
        if ('' != config.get('cols_out', vName)) :
            out_columns_names.append(vName)
    
    out_columns_j = {}
    for vName in out_columns_names :
        tName = config.get('cols_out', vName)
        if  tName in in_columns_j :
            out_columns_j[vName] = in_columns_j[tName]
    print('-----------------------------------')
    for vName in out_columns_j :
        print(vName, '\t', out_columns_j[vName])    
    print('-----------------------------------')
    strHeader = ','.join(out_columns_names) +',группа,подгруппа,бренд,'
    print('HEAD =', strHeader)

    # считываем имена файлов и имя листа
    FilenameIn   = config.get('input','Filename_in' )
    SheetName    = config.get('input','SheetName'   )      
    FilenameOut  = config.get('input','Filename_out')
    print('SHEET=', SheetName)
    
    # считываем признаки группы и подгруппы
    if ('' != config.get('grp_properties',  'группа')) :
        colGrp               = config.getint('grp_properties',     'группа')
    if ('' != config.get('grp_properties',  'подгруппа')) :
        colSGrp              = config.getint('grp_properties',  'подгруппа')
    if ('' != config.get('grp_properties',  'GrpFonti')) :
        GrpFonti             = config.getint('grp_properties',   'GrpFonti')
    if ('' != config.get('grp_properties',  'SubGrpFonti')) :
        SubGrpFonti          = config.getint('grp_properties','SubGrpFonti')
    if ('' != config.get('grp_properties',  'BrandFonti')) :
        BrandFonti           = config.getint('grp_properties', 'BrandFonti')
    if ('' != config.get('grp_properties',  'HeaderFonti')) :
        HeaderFonti          = config.getint('grp_properties','HeaderFonti')
    if ('' != config.get('grp_properties',  'HeaderFontSize')) :
        HeaderFontSize       = config.getint('grp_properties','HeaderFontSize')
    if ('' != config.get('grp_properties',  'RegularFontSize')) :
        RegularFontSize      = config.getint('grp_properties','RegularFontSize')
    if ('' != config.get('grp_properties',  'SubGrpFontSize')): 
        SubGrpFontSize       = config.getint('grp_properties','SubGrpFontSize')
    if ('' != config.get('grp_properties',  'GrpFontSize')) :
        GrpFontSize          = config.getint('grp_properties',   'GrpFontSize')
    if ('' != config.get('grp_properties',  'SubGrpBackgroundColor')) :
        SubGrpBackgroundColor= config.getint('grp_properties','SubGrpBackgroundColor')
    if ('' != config.get('grp_properties',  'GrpBackgroundColor')) :
        GrpBackgroundColor   = config.getint('grp_properties',   'GrpBackgroundColor')
    subgrpfontbold           = config.get('grp_properties','subgrpfontbold')
    grpfontbold              = config.get('grp_properties',   'grpfontbold')
    return FilenameIn



def quoted(sss):
    if ((',' in sss) or ('"' in sss) or ('\n' in sss))  and not(sss[0]=='"' and sss[-1]=='"') :
        sss = '"'+sss.replace('"','""')+'"'
    return sss
